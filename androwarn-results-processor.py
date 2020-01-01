# For parsing html
import os
import re
from bs4 import BeautifulSoup
from bs4 import SoupStrainer
# For storing results
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Create the Workbook for storage later
workbook = Workbook()
worksheet = workbook.active
worksheet.title = "Androwarn"

# Strainer for finding html <body> tag
only_body = SoupStrainer('body')
# Move to listed directory of html files
os.chdir('C:/Users/Casey Dixon/PycharmProjects/androwarn-results-processor/html files')
# Set intial position in excel document
row = 1
column = 1
count = 1
for subdirectory, directory, files in os.walk(os.getcwd()):
    for file in files:
        if file.endswith(".html"):
            print(count)
            try:
                androwarnHTML = BeautifulSoup(open(file), features="html.parser", parse_only=only_body)

                # Find the number of results we need to create
                resultList = androwarnHTML.find_all('ul', attrs={"class": "nav nav-list"})
                liList = re.findall("<li.*", str(resultList))
                iteration = 0
                lowerBound = 0
                upperBound = 0
                # Determine the number of results in the html document from the security section
                for liTag in liList:
                    iteration = iteration + 1
                    if liTag == '<li class="nav-header">Analysis Results</li>':
                        lowerBound = iteration
                    if liTag == '<li class="nav-header">Apk File</li>':
                        upperBound = iteration
                upperIndex = upperBound - lowerBound + 3

                # Find the name of our apk for storing data under
                titleOfApk = androwarnHTML.find_all('div', attrs={"id": "file-name"})
                titleOfApk = re.findall("<h3>(.*?)<*/*h3>", str(titleOfApk))
                worksheet.cell(row=row, column=column).value = titleOfApk[0]
                columnLetter = get_column_letter(column)
                worksheet.column_dimensions[columnLetter].width = 50
                column = column + 1

                # Create initial result list based on div tags
                analysisResults = androwarnHTML.find_all('div', attrs={"class": "tab-pane"})
                resultIndex = 0
                excelH3Tag = ""

                # Adding the counts of each error type
                worksheet.cell(row=row, column=column).value = 'Number of security flaws'
                columnLetter = get_column_letter(column)
                worksheet.column_dimensions[columnLetter].width = 30
                column = column + 1
                worksheet.cell(row=row, column=column).value = upperIndex-4
                column = column + 1

                # Main process for writing to each row
                for liTag in analysisResults:
                    if 3 < resultIndex < upperIndex:
                        h2tag = re.findall("<h2>(.*?)<*/*h2>", str(liTag))
                        h3tag = re.findall("<h3>(.*?)<*/*h3>", str(liTag))
                        worksheet.cell(row=row, column=column).value = h2tag[0]
                        columnLetter = get_column_letter(column)
                        worksheet.column_dimensions[columnLetter].width = 30
                        column = column + 1
                        worksheet.cell(row=row, column=column).value = 'Number of specific flaws'
                        columnLetter = get_column_letter(column)
                        worksheet.column_dimensions[columnLetter].width = 30
                        column = column + 1
                        worksheet.cell(row=row, column=column).value = len(h3tag)
                        columnLetter = get_column_letter(column)
                        worksheet.column_dimensions[columnLetter].width = 5
                        column = column + 1
                        for i in range(len(h3tag)):
                            excelH3Tag = excelH3Tag + h3tag[i] + '\n'
                        worksheet.cell(row=row, column=column).value = excelH3Tag
                        columnLetter = get_column_letter(column)
                        worksheet.column_dimensions[columnLetter].width = 10
                        column = column + 1
                        excelH3Tag = ""
                    resultIndex = resultIndex + 1
                # Increment row reset column for next result
                row = row + 1
                column = 1
                workbook.save("Androwarn Results.xlsx")

                count = count + 1
                if count == 10:
                    break;
            except:
                print(file)
